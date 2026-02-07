"""
Excel document processor.
Extracts text content from Excel spreadsheets.
"""
import io
from typing import Optional
from openpyxl import load_workbook
from .base import BaseProcessor


class ExcelProcessor(BaseProcessor):
    """Processor for Excel documents (XLSX, XLS)"""
    
    def extract_text(self, file_content: bytes, filename: str) -> Optional[str]:
        """
        Extract text from Excel file.
        
        Args:
            file_content: Excel file content as bytes
            filename: Original filename
            
        Returns:
            Extracted text content, or None if extraction fails
        """
        try:
            excel_file = io.BytesIO(file_content)
            workbook = load_workbook(excel_file, data_only=True)
            
            text_parts = []
            
            # Extract text from each worksheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"[Sheet: {sheet_name}]")
                
                # Extract cell values
                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None and str(cell).strip():
                            row_text.append(str(cell).strip())
                    
                    if row_text:
                        text_parts.append(" | ".join(row_text))
            
            if not text_parts:
                return None
            
            full_text = "\n".join(text_parts)
            return self.clean_text(full_text)
        
        except Exception as e:
            print(f"Error processing Excel {filename}: {e}")
            return None
